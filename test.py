from openpyxl import load_workbook


fn = 'Analytics_sending_app.xlsx'
wb = load_workbook(fn)
ws = wb['analytics data']

print(ws['E4'].value)





