from openpyxl import load_workbook
from config import CON, MPKind
#------------------------------------------------------------------------------------------

fn = 'Analytics_sending_app.xlsx'
wb = load_workbook(fn)
ws = wb['current sheet']

#считывание значений границ
f = open('grand.txt')
bord = []
for line in f:
    bord.append(int(line))
f.close()
last_bord = bord[-1]
pred_last_bord = bord[-2]

cod_MPK=[]
item = 2
while ws[f'D{item}'].value is not None: item += 1
index_last_MPK = item

for i in range(2, index_last_MPK):
    cod_MPK.append(ws[f'D{i}'].value)

l = []
cod_MPK_sort = []
for i in cod_MPK:
    if i[:4:] in CON:
        l = i.split('/')
        l = l[0]+'/00'
        cod_MPK_sort.append(l)
    else:
        cod_MPK_sort.append(i[:4:])

cod_MPK_count = {}
for i in cod_MPK_sort:
    cod_MPK_count[f'{i}'] = cod_MPK_sort.count(i)

cod_MPK_count_sort = {}
for key, value in cod_MPK_count.items():
    if key not in cod_MPK_count_sort:
        cod_MPK_count_sort[key] = value

ws = wb['analytics data']
ld = []
for i in range(5, 780):
    c = ws[f'D{i}']
    if c.value is not None: ld.append(c.value)

# for key, value in cod_MPK_count_sort.items():
#     print(f'  {key} : {value}')
#     print('\n')
#     sum_ += value
# print(f'    колличество заявкок - {sum_}')
#----------------------------------------------------------------------------------------------------------------

ad = []
kol = 0

for key, value in cod_MPK_count_sort.items():
    for i in range(5, 780):
        if ws[f'D{i}'].value is not None:
            a = ws[f'D{i}'].value
            b = ws[f'E{i}'].value
            if key == a:
                if b is not None:
                    ws[f'E{i}'] = f'{int(b) + int(value)}'
                    ws[f'A{i}'] = '...'
                ws[f'E{i}'] = value
                ws[f'A{i}'] = '...'
            if key not in ld:
                if key not in ad:
                    ad.append(key)

for i in range(len(bord)):
    if bord[i] == pred_last_bord:
        for j in range(pred_last_bord, last_bord):
            if ws[f'E{j}'].value is not None:
                ws[f'A{bord[i]}'] = '...'
        break
    for j in range(bord[i]+1, bord[i+1]):
        if ws[f'E{j}'].value is not None:
            ws[f'A{bord[i]}'] = '...'

wb.save(fn)
wb.close()

sum_ = 0
sum__ = 0
for value in cod_MPK_count_sort.values():
    sum_ += value

print(f'    колличество заявок - {sum_}\n')
print('МПК которых нет в таблице:')
print(f'{len(ad)} штук')
for i in ad:
    print(i)


